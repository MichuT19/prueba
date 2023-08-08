import csv,openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl import Workbook
from home.cliente.models import Calificacion
from openpyxl.styles import PatternFill
# #Pais
def reportexcel_pais(modeladmin,request,queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reportepaises.xlsx"'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    #Escribir los encabezados
    worksheet['A1'] = 'Id pais'
    worksheet['B1'] = 'Nombre'

    #Escribir los datos de los países
    paises = queryset.values_list('id_pais', 'nombre')
    for row_index, pais in enumerate(paises, start=2):
        for col_index, value in enumerate(pais, start=1):
            column_letter = get_column_letter(col_index)
            cell = worksheet['{}{}'.format(column_letter, row_index)]
            cell.value = value

    workbook.save(response)
    return response
reportexcel_pais.short_description='Reporte en Excel'

def reportexcel_calificacion(modeladmin,request,queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reportecalificaciones.xlsx"'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    #Escribir los encabezados
    worksheet['A1'] = 'ID Paciente'
    worksheet['B1'] = 'ID Trabajador'
    worksheet['C1'] = 'Puntuación'
    worksheet['D1'] = 'Comentario'

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_index, header in enumerate([
        'ID Paciente', 'ID Trabajador','Puntuación','Comentario'], start=1):
        column_letter = get_column_letter(col_index)
        cell = worksheet['{}1'.format(column_letter)]
        cell.value = header
        cell.fill = yellow_fill

    #Escribir los datos de los países
    calificaciones = queryset.values_list('id_paciente__nombre', 'id_trabajador__id_cliente__nombre', 'puntuacion', 'comentario')
    for row_index, calificacion in enumerate(calificaciones, start=2):
        for col_index, value in enumerate(calificacion, start=1):
            column_letter = get_column_letter(col_index)
            cell = worksheet['{}{}'.format(column_letter, row_index)]
            cell.value = value

    workbook.save(response)
    return response
reportexcel_calificacion.short_description = 'Reporte en Excel'

def reportexcel_trabajador(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reportetrabajadores.xlsx"'
    workbook = Workbook()
    worksheet = workbook.active

    # Escribir los encabezados  
    worksheet['A1'] = 'ID Trabajador'
    worksheet['B1'] = 'ID Cliente'
    worksheet['C1'] = 'ID Tipo Trabajador'
    worksheet['D1'] = 'PDF Cédula'
    worksheet['E1'] = 'PDF Curriculum'
    worksheet['F1'] = 'Latitud'
    worksheet['G1'] = 'Longitud'

    # Escribir los datos de los trabajadores
    trabajadores = queryset.values_list('id_trabajador', 'id_cliente__nombre', 'id_tipo_trabajador__descripcion',
                                        'pdf_cedula', 'pdf_curriculum', 'latitud', 'longitud')
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_index, header in enumerate([
        'ID Trabajador','ID Cliente','ID Tipo Trabajador','PDF Cédula','PDF Curriculum',
        'Latitud','Longitud'], start=1):
        column_letter = get_column_letter(col_index)
        cell = worksheet['{}1'.format(column_letter)]
        cell.value = header
        cell.fill = yellow_fill

    for row_index, trabajador in enumerate(trabajadores, start=2):
        for col_index, value in enumerate(trabajador, start=1):
            column_letter = get_column_letter(col_index)
            cell = worksheet['{}{}'.format(column_letter, row_index)]
            cell.value = value

    workbook.save(response)
    return response
reportexcel_trabajador.short_description = 'Reporte en Excel'

def reportexcel_clasificaciones(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporteclasificaciones_enfermedades.xlsx"'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Escribir los encabezados
    worksheet['A1'] = 'ID Clasificación'
    worksheet['B1'] = 'Descripción'

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_index, header in enumerate([
        'ID Clasificación','Descripción'], start=1):
        column_letter = get_column_letter(col_index)
        cell = worksheet['{}1'.format(column_letter)]
        cell.value = header
        cell.fill = yellow_fill    
    

    # Escribir los datos de las clasificaciones
    clasificaciones = queryset.values_list('idclasificacion', 'descripcion')
    for row_index, clasificacion in enumerate(clasificaciones, start=2):
        for col_index, value in enumerate(clasificacion, start=1):
            column_letter = get_column_letter(col_index)
            cell = worksheet['{}{}'.format(column_letter, row_index)]
            cell.value = value

    workbook.save(response)
    return response
reportexcel_clasificaciones.short_description = 'Reporte en Excel'

def reportexcel_cliente(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporteclientes.xlsx"'
    workbook = Workbook()
    worksheet = workbook.active

    # Escribir los encabezados
    worksheet['A1'] = 'ID cliente'
    worksheet['B1'] = 'Cédula'
    worksheet['C1'] = 'Nombre'
    worksheet['D1'] = 'Apellido'
    worksheet['E1'] = 'Fecha de Nacimiento'
    worksheet['F1'] = 'Teléfono'
    worksheet['G1'] = 'Sexo'
    worksheet['H1'] = 'País'
    worksheet['I1'] = 'Provincia'
    worksheet['J1'] = 'Ciudad'
    worksheet['K1'] = 'Referencia de Domicilio'
    worksheet['L1'] = 'Tipo de Sangre'

    # Escribir los datos de los clientes
    clientes = queryset.values_list(
        'id_cliente', 'cedula', 'nombre', 'apellido', 'fecha_nacimiento',
        'telefono', 'sexo__descripcion', 'pais__nombre', 'provincia__nombre',
                                    'ciudad__nombre', 'referencia_de_domicilio', 'tipo_sangre__descripcion'
    )
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_index, header in enumerate([
        'ID cliente', 'Cédula','Nombre','Apellido','Fecha de Nacimiento',
        'Teléfono','Sexo','País','Provincia','Ciudad','Referencia de Domicilio','Tipo de Sangre'], start=1):
        column_letter = get_column_letter(col_index)
        cell = worksheet['{}1'.format(column_letter)]
        cell.value = header
        cell.fill = yellow_fill

    for row_index, cliente in enumerate(clientes, start=2):
        for col_index, value in enumerate(cliente, start=1):
            column_letter = get_column_letter(col_index)
            cell = worksheet['{}{}'.format(column_letter, row_index)]
            cell.value = value

    workbook.save(response)
    return response
reportexcel_cliente.short_description = 'Reporte en Excel'

def reportexcel_enfermedades(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporteenfermedades.xlsx"'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Escribir los encabezados
    worksheet['A1'] = 'ID Enfermedad'
    worksheet['B1'] = 'ID Clasificación Enfermedad'
    worksheet['C1'] = 'Descripción'
    worksheet['D1'] = 'Estado'

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_index, header in enumerate([
        'ID Enfermedad', 'ID Clasificación Enfermedad','Descripción','Estado'], start=1):
        column_letter = get_column_letter(col_index)
        cell = worksheet['{}1'.format(column_letter)]
        cell.value = header
        cell.fill = yellow_fill

    # Escribir los datos de las enfermedades
    enfermedades = queryset.values_list('id_enfermedad', 'id_clasificacionenfermedad__descripcion', 'descripcion', 'estado')
    for row_index, enfermedad in enumerate(enfermedades, start=2):
        for col_index, value in enumerate(enfermedad, start=1):
            column_letter = get_column_letter(col_index)
            cell = worksheet['{}{}'.format(column_letter, row_index)]
            cell.value = value

    workbook.save(response)
    return response
reportexcel_enfermedades.short_description = 'Reporte en Excel'

def reportexcel_enfermedades_paciente(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporteenfermedades_pacientes.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active

    # Escribir los encabezados
    worksheet['A1'] = 'Id Enfermedad'
    worksheet['B1'] = 'Id Cliente'
    worksheet['C1'] = 'Descripción'
    worksheet['D1'] = 'Estado'

    # Obtener los datos de las enfermedades por paciente
    enfermedades_pacientes = queryset.values_list('id_enfermedad__descripcion', 'id_cliente__nombre', 'descripcion', 'estado')

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_index, header in enumerate([
        'ID Enfermedad', 'Id Cliente','Descripción','Estado'], start=1):
        column_letter = get_column_letter(col_index)
        cell = worksheet['{}1'.format(column_letter)]
        cell.value = header
        cell.fill = yellow_fill
        
    # Escribir los datos en las filas
    for row_index, enfermedad_paciente in enumerate(enfermedades_pacientes, start=2):
        for col_index, value in enumerate(enfermedad_paciente, start=1):
            column_letter = get_column_letter(col_index)
            cell = worksheet['{}{}'.format(column_letter, row_index)]
            cell.value = value

    workbook.save(response)
    return response
reportexcel_enfermedades_paciente.short_description = 'Reporte en Excel'

def reportexcel_profesiones(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporteprofesiones.xlsx"'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Escribir los encabezados
    worksheet['A1'] = 'Id profesiones'
    worksheet['B1'] = 'Descripción'
    worksheet['C1'] = 'Estado'

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_index, header in enumerate([
        'Id profesiones', 'Descripción','Estado'], start=1):
        column_letter = get_column_letter(col_index)
        cell = worksheet['{}1'.format(column_letter)]
        cell.value = header
        cell.fill = yellow_fill

    # Escribir los datos de las profesiones
    profesiones = queryset.values_list('id_profesiones', 'descripcion', 'estado')
    for row_index, profesion in enumerate(profesiones, start=2):
        for col_index, value in enumerate(profesion, start=1):
            column_letter = get_column_letter(col_index)
            cell = worksheet['{}{}'.format(column_letter, row_index)]
            cell.value = value

    workbook.save(response)
    return response
reportexcel_profesiones.short_description = 'Reporte en Excel'

def reportexcel_servicio(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporteservicios.xlsx"'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Escribir los encabezados
    worksheet['A1'] = 'ID Servicio'
    worksheet['B1'] = 'ID Profesiones'
    worksheet['C1'] = 'Descripción'
    worksheet['D1'] = 'Estado'
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_index, header in enumerate([
        'ID Servicio', 'ID Profesiones','Descripción','Estado'], start=1):
        column_letter = get_column_letter(col_index)
        cell = worksheet['{}1'.format(column_letter)]
        cell.value = header
        cell.fill = yellow_fill

    # Escribir los datos de los servicios
    servicios = queryset.values_list('id_servicio', 'id_profesiones__descripcion', 'descripcion', 'estado')
    for row_index, servicio in enumerate(servicios, start=2):
        for col_index, value in enumerate(servicio, start=1):
            column_letter = get_column_letter(col_index)
            cell = worksheet['{}{}'.format(column_letter, row_index)]
            cell.value = value

    workbook.save(response)
    return response
reportexcel_servicio.short_description = 'Reporte en Excel'

def reportexcel_cita(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reportecitas.xlsx"'
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Escribir los encabezados
    worksheet['A1'] = 'ID Cita'
    worksheet['B1'] = 'Trabajador'
    worksheet['C1'] = 'Cliente'
    worksheet['D1'] = 'Motivo de Cita'
    worksheet['E1'] = 'Fecha de Creación'
    worksheet['F1'] = 'Fecha Inicio de Atención'
    worksheet['G1'] = 'Fecha Fin de Atención'
    worksheet['H1'] = 'Latitud'
    worksheet['I1'] = 'Longitud'

    # Escribir los datos de las citas
    citas = queryset.values_list(
        'id_cita',
        'id_trabajador__id_cliente__nombre',  # Nombre del campo de clave foránea
        'id_cliente__nombre',  # Nombre del campo de clave foránea
        'descripcion_motivo',
        'fecha_creacion',
        'fecha_inicioatencion',
        'fecha_finatencion',
        'latitud',
        'longitud'
    )

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col_index, header in enumerate([
        'ID Cita', 'Trabajador','Cliente','Motivo de Cita','Fecha de Creación',
        'Fecha Inicio de Atención','Fecha Fin de Atención','Latitud','Longitud'], start=1):
        column_letter = get_column_letter(col_index)
        cell = worksheet['{}1'.format(column_letter)]
        cell.value = header
        cell.fill = yellow_fill


    for row_index, cita in enumerate(citas, start=2):
        for col_index, value in enumerate(cita, start=1):
            column_letter = get_column_letter(col_index)
            cell = worksheet['{}{}'.format(column_letter, row_index)]
            cell.value = value

    workbook.save(response)
    return response

reportexcel_cita.short_description = 'Reporte en Excel'